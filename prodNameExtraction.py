import pandas as pd
from openpyxl import load_workbook
import re

def clean_text(text):
    if not isinstance(text, str):
        return ""
    
    patterns_to_remove = [
        r'ETA[\s\-]*NO\.\s*[\w\-]+',
        r'BIS[\s\-]*NO\.\s*\w+',
        r'DT\.\d{2}\.\d{2}\.\d{2}',
        r'\d{2}\.\d{2}\.\d{2}',
        r'\bR-\d{8,}\b',
        r'CONTENT\s*[:\-]?\s*\d+\.?\d*%',
        r'AVERAGE\s+OF\s+CONTENT\s*\d+\.?\d*%',
        r'\d+\.?\d*%',
    ]
    for pattern in patterns_to_remove:
        text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    
    return re.sub(r'\s+', ' ', text).strip()

def is_model_code(word):
    # Exclude patterns like A362-019-5017-00 or similar model codes
    if bool(re.match(r'^[A-Z0-9\-]{5,}$', word)) and any(char.isdigit() for char in word):
        return True
    # Exclude words with excessive hyphens or numbers
    if word.count('-') > 2 or sum(c.isdigit() for c in word) > 4:
        return True
    return False

def extract_product_name_description(text):
    if not isinstance(text, str) or not text.strip():
        return "", text

    text = clean_text(text)
    text = ' '.join(dict.fromkeys(text.split()))  # Remove exact repeated words

    # If starts with known prefix, keep it
    known_prefixes = [
        "GRID CONNECTED INVERTER",
        "SOLAR INVERTER",
        "GRID TIE SOLAR PV INVERTER",
        "UTILITY INTERCONNECTED PHOTOVOLTAIC INVERTERS",
        "Crystalline Silicon Terrestrial Photovoltaic PV Module",
    ]
    for prefix in known_prefixes:
        if text.upper().startswith(prefix.upper()):
            return prefix.strip(), text

    # Split into segments to avoid model/BIS/ETA pollution
    segments = re.split(r'\b(?:MODEL|ETA|BIS|NO\.|R-|CODE|MAX|INV\.|P\.LIST|BL)\b', text, flags=re.IGNORECASE)
    candidates = []
    for segment in segments:
        words = segment.strip().split()
        clean_words = [w for w in words if not is_model_code(w)]
        candidate = ' '.join(clean_words).strip()
        if 3 <= len(candidate.split()) <= 20:  # Ensure reasonable length
            candidates.append(candidate)

    if candidates:
        # Choose the candidate with the most relevant keywords or longest length
        best = max(candidates, key=lambda x: len(x.split()))
        return best.strip(), text

    return text.strip(), text  # fallback

# === Main Code ===
def process_file(input_path, output_path):
    # Load the workbook and get the sheet names
    wb = load_workbook(input_path, read_only=True)
    sheet = wb.active

    # Read the header row
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    # Ensure 'Product Description' column exists
    if 'Product Description' not in header:
        raise ValueError(f"Column 'Product Description' not found in the uploaded file")

    desc_col_index = header.index('Product Description')

    # Prepare to write the processed file
    processed_chunks = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(header, row))
        product_description = row_dict.get('Product Description', '')
        product_name = extract_product_name_description(product_description)[0]
        row_dict['Product Name'] = product_name
        processed_chunks.append(row_dict)

    # Convert processed chunks to a DataFrame
    processed_df = pd.DataFrame(processed_chunks)

    # Reorder columns to place 'Product Name' before 'Product Description'
    cols = list(processed_df.columns)
    cols.insert(cols.index('Product Description'), cols.pop(cols.index('Product Name')))
    processed_df = processed_df[cols]

    # Save the processed DataFrame to an Excel file
    processed_df.to_excel(output_path, index=False)